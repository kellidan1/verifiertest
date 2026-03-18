import os
import openpyxl
from openpyxl.styles import PatternFill
from source.excel_marksheet_handler import extract_all_students_data, process_student_record

if __name__ == "__main__":
    # Example usage (assuming 'testcl.xlsx' exists in the same directory):
    file_name = "test_data/testcl.xlsx"
    if os.path.exists(file_name):
         print("Reading Excel file...")
         all_students = extract_all_students_data(file_name)    
         print(f"Found {len(all_students)} student records to process.")
         
         # Load the workbook again to modify it
         wb = openpyxl.load_workbook(file_name)
         sheet = wb.active
         
         green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
         red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
         
         # Process all students sequentially
         for student in all_students:
             row, name_var, s1_n, s1_m, s2_n, s2_m, merit_val, perc_val = process_student_record(student)
             
             # Fill Sem 1 marks (col 14)
             if s1_m is True:
                 sheet.cell(row=row, column=14).fill = green_fill
             elif s1_m is False:
                 sheet.cell(row=row, column=14).fill = red_fill
                 
             # Fill Sem 2 marks (col 19)
             if s2_m is True:
                 sheet.cell(row=row, column=19).fill = green_fill
             elif s2_m is False:
                 sheet.cell(row=row, column=19).fill = red_fill
                 
             # Fill Name (col 2)
             # Logic: If any evaluated name match is False => RED. If all evaluated are True => GREEN.
             name_matches = [m for m in (s1_n, s2_n) if m is not None]
             if name_matches:
                 if all(name_matches):
                     sheet.cell(row=row, column=2).fill = green_fill
                 else:
                     sheet.cell(row=row, column=2).fill = red_fill
                     
             # Fill Merit Scholarship (col 6)
             if merit_val is not None and str(merit_val).strip() != "":
                 perc_num = 0.0
                 try:
                     if isinstance(perc_val, str):
                         perc_str = perc_val.replace('%', '').strip()
                         perc_num = float(perc_str)
                         if perc_val.endswith('%'):
                             perc_num = perc_num / 100.0
                         elif perc_num > 1.0:
                             perc_num = perc_num / 100.0
                     elif perc_val is not None:
                         perc_num = float(perc_val)
                         if perc_num > 1.0:
                             perc_num = perc_num / 100.0
                 except Exception:
                     perc_num = 0.0
                     
                 if perc_num >= 0.65:
                     sheet.cell(row=row, column=6).fill = green_fill
                 else:
                     sheet.cell(row=row, column=6).fill = red_fill
                     
         print("\nSequential processing completed!")
         
         os.makedirs("outputs", exist_ok=True)
         output_filename = file_name.replace('.xlsx', 'outputs/_output.xlsx')
         wb.save(output_filename)
         print(f"Saved highlighted verification file as: {output_filename}")
