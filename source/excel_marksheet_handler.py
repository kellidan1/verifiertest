from openpyxl import Workbook
import os
import re
import requests
import openpyxl
import fitz
import pytesseract
from PIL import Image
import io

def extract_all_students_data(file_path: str):
    """
    Reads from row 2 to the end of an Excel file.
    Extracts name (col 2), sem1 marks (col 14), marksheet url 1 (col 17),
    sem2 marks (col 19), and marksheet url 2 (col 22).
    """
    students_data = []
    try:
        # Load the workbook and select the active worksheet
        wb: Workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        if sheet is None:
            raise ValueError("No active sheet found in the workbook.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return students_data

    # Iterate from row 2 to the end
    for row in range(2, sheet.max_row + 1):
        # 1. Go to name (column 2) and extract it
        student_name = sheet.cell(row=row, column=2).value
        if not student_name:
            continue # Skip empty rows
            
        # 2. Go to marks of sem1 (column 14) and extract it
        marks_sem1 = sheet.cell(row=row, column=14).value
        
        # 3. Marksheet link (column 17)
        marksheet_link1 = sheet.cell(row=row, column=17).value
        
        # 4. Go to marks of sem2 (column 19)
        marks_sem2 = sheet.cell(row=row, column=19).value
        
        # 5. Marksheet link 2 (column 22)
        marksheet_link2 = sheet.cell(row=row, column=22).value
        
        # 6. Merit Scholarship (column 6)
        merit_scholarship = sheet.cell(row=row, column=6).value
        
        # 7. Percentage (column 26)
        percentage = sheet.cell(row=row, column=26).value
        
        students_data.append((row, student_name, marks_sem1, marksheet_link1, marks_sem2, marksheet_link2, merit_scholarship, percentage))
        
    return students_data


def download_marksheet(url, save_filename="marksheet"):
    """
    Downloads a file (image or pdf) from the given URL 
    and saves it into a 'temp' directory.
    """
    if not url:
        print("No valid URL provided to download.")
        return None
        
    # Check if it's a Google Drive link and convert to direct download
    if 'drive.google.com' in url:
        match_d = re.search(r'/file/d/([a-zA-Z0-9_-]+)', url)
        if match_d:
            file_id = match_d.group(1)
            url = f"https://drive.google.com/uc?export=download&id={file_id}"
        else:
            match_id = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', url)
            if match_id:
                file_id = match_id.group(1)
                url = f"https://drive.google.com/uc?export=download&id={file_id}"

    folder_name = "temp"
    
    # Create the folder if it doesn't already exist
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        
    try:
        print(f"Downloading from {url} ...")
        response = requests.get(url, stream=True)
        response.raise_for_status()
        
        # Guess extension from Content-Type
        content_type = response.headers.get('content-type', '').lower()
        if 'application/pdf' in content_type:
            ext = 'pdf'
        elif 'image/jpeg' in content_type or 'image/jpg' in content_type:
            ext = 'jpg'
        elif 'image/png' in content_type:
            ext = 'png'
        else:
            # Fallback to the URL ending
            ext = url.split('.')[-1][:4]
            if not ext.isalnum():
                ext = 'pdf' # default extension
                
        # Construct the full path to save the file
        file_path = os.path.join(folder_name, f"{save_filename}.{ext}")
        
        # Write the downloaded content to the file in chunks
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    
        print(f"Successfully saved to {file_path}")
        return file_path
        
    except requests.exceptions.RequestException as e:
        print(f"Error downloading the file: {e}")
        return None


def extract_text_with_ocr(file_path):
    """
    Extracts text from a given image or PDF file using OCR.
    """
    text = ""
    try:
        # Check if it's a PDF
        if file_path.lower().endswith('.pdf'):
            doc = fitz.open(file_path)
            for page in doc:
                pix = page.get_pixmap(dpi=300)
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                text += pytesseract.image_to_string(img)
            doc.close()
        # Otherwise assume it's an image
        elif file_path.lower().endswith(('.png', '.jpg', '.jpeg')):
            img = Image.open(file_path)
            text += pytesseract.image_to_string(img)
        else:
            print(f"Unsupported file type for OCR: {file_path}")
            
    except Exception as e:
        print(f"Error extracting text with OCR: {e}")
        
    return text


def verify_marksheet_data(ocr_text, expected_name, expected_marks):
    """
    Compares the expected name and marks with the text extracted via OCR.
    Returns a dictionary with match results.
    """
    results = {
        "name_match": False,
        "marks_match": False,
        "ocr_marks": None
    }
    
    # Normalize strings for comparison (remove extra spaces, case-insensitive)
    def normalize_str(s):
        return re.sub(r'\s+', ' ', str(s)).strip().upper()
        
    ocr_text_normalized = normalize_str(ocr_text)
    expected_name_normalized = normalize_str(expected_name)
    expected_marks_str = str(expected_marks).strip()
    
    # Check if the exact expected name exists in the OCR text sequentially
    if expected_name_normalized in ocr_text_normalized:
        results["name_match"] = True
            
    # Check if the marks exist in the OCR text
    if expected_marks_str in ocr_text_normalized:
        results["marks_match"] = True
        results["ocr_marks"] = expected_marks_str
    else:
        # Attempt to extract actual marks using regex near GRAND TOTAL
        match = re.search(r'GRAND TOTAL.*?(\d{2,4})', ocr_text_normalized)
        if match:
            results["ocr_marks"] = match.group(1)
        else:
            # Fallback looking for word TOTAL
            match2 = re.search(r'TOTAL.*?(\d{2,4})', ocr_text_normalized)
            if match2:
                results["ocr_marks"] = match2.group(1)
            else:
                results["ocr_marks"] = "Not Found"
        
    return results


def process_student_record(student_data):
    """
    Processes a single student's record: downloads both marksheets, runs OCR, and verifies.
    Returns (row, name, s1_n, s1_m, s2_n, s2_m, merit, percentage).
    """
    row, name_var, marks_sem1_var, link1_var, marks_sem2_var, link2_var, merit_scholarship, percentage = student_data
    print(f"\n--- Processing: {name_var} ---")
    
    sem1_name_match = None
    sem1_marks_match = None
    sem2_name_match = None
    sem2_marks_match = None
    
    # --- SEMESTER 1 ---
    if link1_var:
        print(f"[{name_var}] Sem 1 Expected: {marks_sem1_var}")
        safe_name = "".join(c for c in str(name_var) if c.isalnum() or c in "_-").strip()
        downloaded_file1 = download_marksheet(link1_var, save_filename=f"{safe_name}_sem1_marksheet")
        
        if downloaded_file1:
            print(f"[{name_var}] Performing OCR on Sem 1 marksheet...")
            ocr_text1 = extract_text_with_ocr(downloaded_file1)
            
            print(f"[{name_var}] Verifying Sem 1 data...")
            res1 = verify_marksheet_data(ocr_text1, name_var, marks_sem1_var)
            
            print(f"[{name_var}] Sem 1 Name Match: {'YES' if res1['name_match'] else 'NO'}")
            print(f"[{name_var}] Sem 1 Marks Match: {'YES' if res1['marks_match'] else 'NO'}")
            if not res1['marks_match']:
                print(f"[{name_var}] -> Expected: {marks_sem1_var}, OCR Found: {res1['ocr_marks']}")
            
            sem1_name_match = res1['name_match']
            sem1_marks_match = res1['marks_match']
            
            if sem1_name_match and sem1_marks_match:
                print(f"*** [{name_var}] SEM 1 VERIFICATION SUCCESSFUL! ***")
            else:
                print(f"*** [{name_var}] SEM 1 VERIFICATION FAILED! ***")
    else:
        print(f"[{name_var}] Skipping Sem 1: No marksheet link provided.")
        
    print("-" * 30)
        
    # --- SEMESTER 2 ---
    if link2_var:
        print(f"[{name_var}] Sem 2 Expected: {marks_sem2_var}")
        safe_name = "".join(c for c in str(name_var) if c.isalnum() or c in "_-").strip()
        downloaded_file2 = download_marksheet(link2_var, save_filename=f"{safe_name}_sem2_marksheet")
        
        if downloaded_file2:
            print(f"[{name_var}] Performing OCR on Sem 2 marksheet...")
            ocr_text2 = extract_text_with_ocr(downloaded_file2)
            
            print(f"[{name_var}] Verifying Sem 2 data...")
            res2 = verify_marksheet_data(ocr_text2, name_var, marks_sem2_var)
            
            print(f"[{name_var}] Sem 2 Name Match: {'YES' if res2['name_match'] else 'NO'}")
            print(f"[{name_var}] Sem 2 Marks Match: {'YES' if res2['marks_match'] else 'NO'}")
            if not res2['marks_match']:
                print(f"[{name_var}] -> Expected: {marks_sem2_var}, OCR Found: {res2['ocr_marks']}")
            
            sem2_name_match = res2['name_match']
            sem2_marks_match = res2['marks_match']
            
            if sem2_name_match and sem2_marks_match:
                print(f"*** [{name_var}] SEM 2 VERIFICATION SUCCESSFUL! ***")
            else:
                print(f"*** [{name_var}] SEM 2 VERIFICATION FAILED! ***")
    else:
        print(f"[{name_var}] Skipping Sem 2: No marksheet link provided.")
        
    return row, name_var, sem1_name_match, sem1_marks_match, sem2_name_match, sem2_marks_match, merit_scholarship, percentage