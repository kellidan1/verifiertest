from source.common.constants import ERROR_COLOR, SUCCESS_COLOR
from openpyxl.worksheet.worksheet import Worksheet
import fitz
import io
import numpy as np
from paddleocr import PaddleOCR
from PIL import Image
import re
from openpyxl.styles import PatternFill


def _check_fee_match(
    ocr: PaddleOCR, sheet: Worksheet, pdf_path: str, row: int, col: int
) -> bool:

    # TODO: Add Support for IMG/PNG Files

    cell_value = sheet.cell(row, col).value

    if not cell_value:
        return False

    # Clean Excel value (remove commas, etc.)
    cell_digits = re.sub(r"[^\d]", "", str(cell_value))

    if not cell_digits:
        return False

    if not pdf_path.endswith(".pdf"):
        # TODO: Throw UnImplemented File type Error
        return False

    # Open PDF
    doc = fitz.open(pdf_path)

    # Check all pages
    for i in range(len(doc)):
        page = doc.load_page(i)
        pix = page.get_pixmap(dpi=300)
        image = Image.open(io.BytesIO(pix.tobytes("png")))

        # OCR
        results = ocr.ocr(np.array(image), cls=True)

        if results and results[0]:
            full_content = " ".join([line[1][0] for line in results[0]])
            full_content = re.sub(r" +", " ", full_content)

            # Flexible match (handles spacing issues)
            pattern = r"(INR|Rs\.?)\s*[:\-]?\s*" + r"\s*".join(list(cell_digits))

            if re.search(pattern, full_content, re.IGNORECASE):
                doc.close()
                return True

    doc.close()
    return False


def fee_reciept_check(
    ocr: PaddleOCR, sheet: Worksheet | None, col: int = -1
):

    if sheet is None or col == -1:
        return -1

    green_fill = PatternFill(
        start_color=SUCCESS_COLOR, end_color=SUCCESS_COLOR, fill_type="solid"
    )
    red_fill = PatternFill(
        start_color=ERROR_COLOR, end_color=ERROR_COLOR, fill_type="solid"
    )

    # TODO: handle Batch Parallel processing for faster completion, Add upper limit
    for row in range(2, sheet.max_row + 1):
        if not sheet.cell(row=row, column=2).value:
            continue
        result = _check_fee_match(ocr, sheet, f"temp/fees/row{row}.pdf", row, col)
        if result:
            sheet.cell(row=row, column=col).fill = green_fill
        else:
            sheet.cell(row=row, column=col).fill = red_fill
    return 0
