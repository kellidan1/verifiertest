import openpyxl
from openpyxl import Workbook

def extract_all_students_data(file_path: str):
    students_data = []
    
    try:
        wb: Workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return students_data

    # Get headers for col 17 and 22
    def sanitize(s):
        if not s: return ""
        return "".join(c for c in str(s) if c.isalnum() or c in "_-").strip()
    
    h1 = sanitize(sheet.cell(row=1, column=17).value) or "Col17"
    h2 = sanitize(sheet.cell(row=1, column=22).value) or "Col22"

    for row in range(2, sheet.max_row + 1):
        student_name = sheet.cell(row=row, column=2).value
        if not student_name:
            continue

        students_data.append((
            row,
            student_name,
            sheet.cell(row=row, column=14).value,
            sheet.cell(row=row, column=17).value,
            sheet.cell(row=row, column=19).value,
            sheet.cell(row=row, column=22).value,
            sheet.cell(row=row, column=6).value,
            sheet.cell(row=row, column=26).value,
            h1,
            h2
        ))

    return students_data