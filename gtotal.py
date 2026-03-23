from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def is_red(cell):
    color = cell.fill.fgColor.rgb
    return color is not None and "FF0000" in color


def check_gtotal(sheet, row, green_fill, red_fill):
    try:
        cell14 = sheet.cell(row=row, column=14)
        cell19 = sheet.cell(row=row, column=19)
        cell15 = sheet.cell(row=row, column=15)
        cell20 = sheet.cell(row=row, column=20)
        result_cell = sheet.cell(row=row, column=26)

        # 🔴 RULE: if ANY source cell is red → result must be red
        if (is_red(cell14) or is_red(cell19) or 
            is_red(cell15) or is_red(cell20)):
            result_cell.fill = red_fill
            return

        val14 = cell14.value or 0
        val19 = cell19.value or 0
        val15 = cell15.value or 0
        val20 = cell20.value or 0

        grandtotal = int(val14) + int(val19)
        maxmarks = int(val15) + int(val20)

        if maxmarks == 0:
            result_cell.fill = red_fill
            return

        calculated_percentage = round((grandtotal / maxmarks) * 100, 2)
        actual_value = result_cell.value

        if actual_value is not None and abs(calculated_percentage - float(actual_value)) < 0.01:
            result_cell.fill = green_fill
        else:
            result_cell.fill = red_fill

    except Exception as e:
        print(f"Error in row {row}: {e}")


def check_total(sheet, row, green_fill, red_fill):
    try:
        cell14 = sheet.cell(row=row, column=14)
        cell19 = sheet.cell(row=row, column=19)
        result_cell = sheet.cell(row=row, column=24)

        # 🔴 RULE: if either input is red → result must be red
        if is_red(cell14) or is_red(cell19):
            result_cell.fill = red_fill
            return

        val14 = cell14.value or 0
        val19 = cell19.value or 0

        total = int(val14) + int(val19)
        actual_value = result_cell.value

        if actual_value is not None and total == int(actual_value):
            result_cell.fill = green_fill
        else:
            result_cell.fill = red_fill

    except Exception as e:
        print(f"Error in row {row}: {e}")

# --- Main execution ---
file_path = "outputs/testcl_output.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active  

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for row in range(2, sheet.max_row + 1):
    if not sheet.cell(row=row, column=2).value:
        continue
    check_total(sheet, row, green_fill, red_fill)
    check_gtotal(sheet, row, green_fill, red_fill)

workbook.save(file_path)