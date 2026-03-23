from openpyxl import load_workbook
from openpyxl.styles import PatternFill

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def is_red(cell):
    fill = cell.fill

    if fill is None or fill.fill_type != "solid":
        return False

    color = fill.start_color

    if color is None or color.rgb is None:
        return False

    return color.rgb[-6:] == "FF0000"
    


# Load the Excel file
file_path = "outputs/testcl_output.xlsx"  # replace with your file path
workbook = load_workbook(file_path)

# Select a sheet
sheet = workbook.active  # or use workbook["Sheet1"]

#test logic
test_cell = sheet.cell(row=8, column=2)

print("Cell value:", test_cell.value)
print("RGB color:", test_cell.fill.start_color.rgb)
print("Is red:", is_red(test_cell))


