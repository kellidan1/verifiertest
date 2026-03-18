import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import PatternFill

# Import backend logic
from excel_marksheet_handler import extract_all_students_data, process_student_record

class VerifierApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Marksheet Verifier Tool")
        self.root.geometry("500x250")
        self.root.resizable(False, False)
        
        self.file_path_var = tk.StringVar()
        
        self.create_widgets()
        
    def create_widgets(self):
        # File Selection Frame
        frame_top = tk.Frame(self.root, pady=20)
        frame_top.pack(fill=tk.X, padx=20)
        
        tk.Label(frame_top, text="Select Excel File:").pack(anchor=tk.W)
        
        entry_file = tk.Entry(frame_top, textvariable=self.file_path_var, width=50)
        entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        btn_browse = tk.Button(frame_top, text="Browse...", command=self.browse_file)
        btn_browse.pack(side=tk.RIGHT)
        
        # Verification Button
        self.btn_verify = tk.Button(self.root, text="Verify Data", command=self.start_verification, width=20, height=2, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
        self.btn_verify.pack(pady=10)
        
        # Progress Bar and Status
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=20, pady=10)
        
        self.status_label = tk.Label(self.root, text="Ready", fg="gray")
        self.status_label.pack()

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if filename:
            self.file_path_var.set(filename)
            
    def start_verification(self):
        file_path = self.file_path_var.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return
            
        self.btn_verify.config(state=tk.DISABLED)
        self.progress_var.set(0)
        self.status_label.config(text="Extracting data from Excel...")
        
        # Start processing in a separate thread so GUI doesn't freeze
        thread = threading.Thread(target=self.run_process, args=(file_path,))
        thread.daemon = True
        thread.start()
        
    def run_process(self, file_path):
        try:
            all_students = extract_all_students_data(file_path)
            total_students = len(all_students)
            
            if total_students == 0:
                self.root.after(0, self.finish_processing, False, "No student records found.")
                return
                
            self.root.after(0, lambda: self.status_label.config(text=f"Found {total_students} records. Beginning OCR..."))
            
            # Load workbook to apply highlights
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            for index, student in enumerate(all_students):
                current_name = student[1]
                self.root.after(0, lambda i=index, n=current_name: self.status_label.config(text=f"Processing {n} ({i+1}/{total_students})..."))
                
                # Actually process the record via OCR
                row, name_var, s1_n, s1_m, s2_n, s2_m, merit_val, perc_val = process_student_record(student)
                
                # Apply fill colors Based on Sem 1
                if s1_m is True:
                    sheet.cell(row=row, column=14).fill = green_fill
                elif s1_m is False:
                    sheet.cell(row=row, column=14).fill = red_fill
                    
                # Based on Sem 2
                if s2_m is True:
                    sheet.cell(row=row, column=19).fill = green_fill
                elif s2_m is False:
                    sheet.cell(row=row, column=19).fill = red_fill
                    
                # Based on Name
                name_matches = [m for m in (s1_n, s2_n) if m is not None]
                if name_matches:
                    if all(name_matches):
                        sheet.cell(row=row, column=2).fill = green_fill
                    else:
                        sheet.cell(row=row, column=2).fill = red_fill
                        
                # Based on Merit Scholarship
                if merit_val is not None and str(merit_val).strip() != "":
                    perc_num = 0.0
                    try:
                        if isinstance(perc_val, str):
                            perc_str = perc_val.replace('%', '').strip()
                            perc_num = float(perc_str)
                            if perc_val.endswith('%'):
                                perc_num = perc_num / 100.0
                            elif perc_num > 1.0:
                                perc_num = perc_num / 100.0
                        elif perc_val is not None:
                            perc_num = float(perc_val)
                            if perc_num > 1.0:
                                perc_num = perc_num / 100.0
                    except Exception:
                        perc_num = 0.0
                        
                    if perc_num >= 0.65:
                        sheet.cell(row=row, column=6).fill = green_fill
                    else:
                        sheet.cell(row=row, column=6).fill = red_fill
                        
                # Update progress bar
                progress_percentage = ((index + 1) / total_students) * 100
                self.root.after(0, self.progress_var.set, progress_percentage)
                
            output_filename = file_path.replace('.xlsx', '_output.xlsx')
            wb.save(output_filename)
            self.root.after(0, self.finish_processing, True, output_filename)
            
        except Exception as e:
            self.root.after(0, self.finish_processing, False, str(e))
            
    def finish_processing(self, success, result):
        self.btn_verify.config(state=tk.NORMAL)
        if success:
            self.status_label.config(text="Verification Complete!")
            self.progress_var.set(100)
            messagebox.showinfo("Success", f"Data verification completed successfully.\nSaved to: {result}\n\nThe output file will now open.")
            
            # Automatically open the generated file using the default OS application
            try:
                os.startfile(result)
            except Exception as e:
                print("Could not open file automatically:", e)
        else:
            self.status_label.config(text="Error occurred")
            messagebox.showerror("Error", f"An error occurred during verification:\n{result}")

if __name__ == "__main__":
    root = tk.Tk()
    app = VerifierApp(root)
    root.mainloop()
